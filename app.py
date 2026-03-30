import streamlit as st
import pandas as pd
from openpyxl import load_workbook
import tempfile

st.title("POA Generator")

uploaded_file = st.file_uploader("Upload Contracts CSV", type=["csv"])

if uploaded_file:
    df = pd.read_csv(uploaded_file)

    # -------------------------------
    # CLEAN COLUMN NAMES
    # -------------------------------
    df.columns = df.columns.str.strip().str.lower().str.replace(" ", "_")

    # -------------------------------
    # EXACT MAPPING (YOUR CSV)
    # -------------------------------
    df = df.rename(columns={
        'vendor': 'tool_name',
        'end_date': 'end_date',
        'total_cost_(usd)': 'contract_value',
        'contract_owner': 'owner'
    })

    # -------------------------------
    # VALIDATION
    # -------------------------------
    required_cols = ['tool_name', 'end_date', 'contract_value']

    missing = [col for col in required_cols if col not in df.columns]

    if missing:
        st.error(f"Missing required columns: {missing}")
        st.write("Your columns are:", df.columns)
        st.stop()

    # -------------------------------
    # CLEAN DATA
    # -------------------------------
    df['contract_value'] = (
        df['contract_value']
        .astype(str)
        .str.replace('$', '', regex=False)
        .str.replace(',', '', regex=False)
    )

    df['contract_value'] = pd.to_numeric(df['contract_value'], errors='coerce')
    df['end_date'] = pd.to_datetime(df['end_date'], errors='coerce')

    df = df.dropna(subset=['contract_value', 'end_date'])

    # -------------------------------
    # QUARTER LOGIC
    # -------------------------------
    def get_quarter(date):
        return f"Q{((date.month - 1) // 3) + 1}"

    df['quarter'] = df['end_date'].apply(get_quarter)

    # -------------------------------
    # SAVINGS (5% - 7%)
    # -------------------------------
    df['low_savings'] = (df['contract_value'] * 0.05).round(2)
    df['high_savings'] = (df['contract_value'] * 0.07).round(2)

    # -------------------------------
    # CONFIDENCE
    # -------------------------------
    def get_confidence(value):
        if value >= 50000:
            return "High"
        elif value >= 10000:
            return "Medium"
        return "Low"

    df['confidence'] = df['contract_value'].apply(get_confidence)

    # -------------------------------
    # SCENARIO
    # -------------------------------
    def get_scenario(value):
        if value >= 50000:
            return "🟡 FLAT"
        else:
            return "🟠 DOWNGRADE"

    # -------------------------------
    # LOAD TEMPLATE
    # -------------------------------
    template_path = "Updated POA Template - Buyers.xlsx"
    wb = load_workbook(template_path)

    # -------------------------------
    # FILL TEMPLATE
    # -------------------------------
    def fill_sheet(sheet_name, data):
        ws = wb[sheet_name]
        row_num = 5

        for _, row in data.iterrows():
            val = row['contract_value']

            ws.cell(row=row_num, column=1).value = row['tool_name']
            ws.cell(row=row_num, column=2).value = row['end_date']
            ws.cell(row=row_num, column=3).value = round(val / 1000, 2)
            ws.cell(row=row_num, column=4).value = get_scenario(val)
            ws.cell(row=row_num, column=5).value = round((val * 0.05) / 1000, 2)
            ws.cell(row=row_num, column=6).value = round((val * 0.07) / 1000, 2)
            ws.cell(row=row_num, column=7).value = row['confidence']

            row_num += 1

    for q in ["Q1", "Q2", "Q3", "Q4"]:
        fill_sheet(q, df[df['quarter'] == q])

    # -------------------------------
    # OUTPUT
    # -------------------------------
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        wb.save(tmp.name)

        st.success("POA generated successfully")

        st.download_button(
            label="Download POA",
            data=open(tmp.name, "rb"),
            file_name="POA_Output.xlsx"
        )
